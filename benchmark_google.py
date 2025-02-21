from google import genai
import openpyxl
from tqdm import tqdm
from def_analysis import get_function_definitions
from contextlib import contextmanager
import signal
from utils import run_test, sanity_check, extract_code
import os
from dotenv import load_dotenv, find_dotenv

load_dotenv(find_dotenv())
client = genai.Client(api_key=os.environ["GOOGLE_KEY"])
model = "gemini-flash-2.0"

@contextmanager
def timeout(seconds):
    def timeout_handler(signum, frame):
        raise TimeoutError(f"Code execution timed out after {seconds} seconds")
    
    # Register the signal handler
    original_handler = signal.signal(signal.SIGALRM, timeout_handler)
    signal.alarm(seconds)
    
    try:
        yield
    finally:
        # Restore the original handler and disable the alarm
        signal.alarm(0)
        signal.signal(signal.SIGALRM, original_handler)

workbook = openpyxl.load_workbook('test.xlsx')
sheet = workbook.active
sanity_check(sheet)

print(f"Running benchmark for model: {model}")
sheet.cell(row=1, column=8, value=model)
correct = 0

for row_idx, row in enumerate(tqdm(sheet.iter_rows(min_row=2, max_row=24, values_only=True)), start=2):
    question, packaged_code, test_cases = row[1:4]
    
    function_defs = "\n".join([
        f"{func.name}({', '.join(func.args)})" 
        for func in get_function_definitions(packaged_code)
    ])

    user_prompt = f"""
    {question}
    
    Output format: Return ONLY the raw function code with no formatting markers, markdown, or XML tags, or ```python.
    Your code should not include any print statements or I/O operations.
    Your code must implement these exact function signatures:
    {function_defs}

    Please include the final code between <code> and </code> tags.
    """

    try:
        response = client.models.generate_content(
            model="gemini-2.0-flash",
            contents=user_prompt)
        
        generated_code = extract_code(response.text)
        print(generated_code)
        # Execute with 5 second timeout
        with timeout(5):
            run_test(generated_code, test_cases)
        sheet.cell(row=row_idx, column=8, value="Passed")
        correct += 1
        
    except TimeoutError:
        sheet.cell(row=row_idx, column=8, value="Failed: Timeout")
    except Exception as e:
        print(e)
        sheet.cell(row=row_idx, column=8, value=f"Failed: {type(e).__name__}")
        
    workbook.save('test.xlsx')

print(f"Model {model}: {correct}/24 correct")
print("=" * 50)

workbook.save('test.xlsx')
print("Benchmarking complete!")