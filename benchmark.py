import aisuite as ai
from tqdm import tqdm
from dotenv import load_dotenv, find_dotenv
import openpyxl
from def_analysis import get_function_definitions
import signal
from contextlib import contextmanager
from utils import sanity_check, run_test, extract_code
from ratelimit import sleep_and_retry, limits

load_dotenv(find_dotenv())

client = ai.Client()
client.configure({
  "together" : {
    "timeout": 1000,
  },
  "openai": {
      "timeout": 1000,
  },
  "google":{
      "timeout": 1000
  }})

workbook = openpyxl.load_workbook('test.xlsx')
sheet = workbook.active
sanity_check(sheet)

@contextmanager
def timeout(seconds):
    def timeout_handler(signum, frame):
        raise TimeoutError(f"Code execution timed out after {seconds} seconds")
    
    # Register the signal handler
    original_handler = signal.signal(signal.SIGALRM, timeout_handler)
    signal.alarm(seconds)
    
    try:
        yield
    finally:
        # Restore the original handler and disable the alarm
        signal.alarm(0)
        signal.signal(signal.SIGALRM, original_handler)

models = [
    "together:meta-llama/Meta-Llama-3.1-405B-Instruct-Turbo",
    "together:deepseek-ai/DeepSeek-R1",
    "openai:o3-mini"
]

# r1 on together tier 1 has 3 reqs/minute
@sleep_and_retry
@limits(calls=3, period=60)
def call_deepseek(message):
    response = client.chat.completions.create(
                model=model, 
                messages=[{"role": "user", "content": message.strip()}])
    generated_code = response.choices[0].message.content.strip()
    return generated_code

for model_idx, model in enumerate(models):
    print(f"Running benchmark for model: {model}")
    sheet.cell(row=1, column=5+model_idx, value=model)
    correct = 0
    
    for row_idx, row in enumerate(tqdm(sheet.iter_rows(min_row=2, max_row=24, values_only=True)), start=2):
        question, packaged_code, test_cases = row[1:4]
        
        function_defs = "\n".join([
            f"{func.name}({', '.join(func.args)})" 
            for func in get_function_definitions(packaged_code)
        ])

        user_prompt = f"""
        {question}
        
        Output format: Return ONLY the raw function code with no formatting markers, markdown, or XML tags, or ```python.
        Your code should not include any print statements or I/O operations.
        Your code must implement these exact function signatures:
        {function_defs}

        Please include the final code between <code> and </code> tags.
        """

        try:
            if model == "together:deepseek-ai/DeepSeek-R1":
                response = call_deepseek(user_prompt)
            else:
                response = client.chat.completions.create(
                    model=model, 
                    messages=[{"role": "user", "content": user_prompt.strip()}]
                )
                
                response = response.choices[0].message.content.strip()
            generated_code = extract_code(response)
            print(generated_code)
            # Execute with 5 second timeout
            with timeout(5):
                run_test(generated_code, test_cases)
            sheet.cell(row=row_idx, column=5+model_idx, value="Passed")
            correct += 1
            
        except TimeoutError:
            sheet.cell(row=row_idx, column=5+model_idx, value="Failed: Timeout")
        except Exception as e:
            print(e)
            sheet.cell(row=row_idx, column=5+model_idx, value=f"Failed: {type(e).__name__}")
            
        workbook.save('test.xlsx')
    
    print(f"Model {model}: {correct}/24 correct")
    print("=" * 50)
    
workbook.save('test.xlsx')
print("Benchmarking complete!")